from __future__ import annotations

from html import escape
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from rich.console import Console
from rich.table import Table

from performance import PERIODS
from table_highlighting import HIGHLIGHT_BOTTOM, HIGHLIGHT_TOP, PerformanceHighlight, build_period_highlights


def _format_value(value: float | None, error: bool = False, highlight: PerformanceHighlight | None = None) -> str:
    if error:
        return "[bold red]Error[/bold red]"
    if value is None:
        return "N/A"
    if highlight == HIGHLIGHT_TOP:
        return f"[bold green]{value * 100:.1f}%[/bold green]"
    if highlight == HIGHLIGHT_BOTTOM:
        return f"[bold red]{value * 100:.1f}%[/bold red]"
    color = "green" if value >= 0 else "red"
    return f"[{color}]{value * 100:.1f}%[/{color}]"


def _format_percent(value: float | None, error: bool = False) -> str:
    if error:
        return "Error"
    if value is None:
        return "N/A"
    return f"{value * 100:.1f}%"


def _format_public_date(latest_date) -> str:
    if latest_date is None:
        return "latest public date"
    timestamp = pd.Timestamp(latest_date)
    return f"{timestamp:%Y-%m-%d}"


def _plain_row_label(row: dict) -> str:
    label = str(row["Fund"])
    if row.get("is_disabled"):
        return f"{label} ({row.get('disabled_reason') or 'Source pending'})"
    if _row_is_stale(row):
        label = f"{label} (as of {_format_public_date(row.get('latest_date'))}, stale {row['stale_days']}d)"
    return label


def _safe_period_value(row: dict, period: str) -> float:
    value = row.get(period)
    return float("-inf") if value is None else float(value)


def _row_is_stale(row: dict) -> bool:
    return bool(row.get("is_stale", row.get("stale_days", 0) > 0))


def _is_firetrail(row: dict) -> bool:
    return str(row.get("Fund", "")).casefold().startswith("firetrail ")


def _average_by_style(rows: list[dict], period: str) -> list[dict[str, object]]:
    grouped: dict[str, list[float]] = {}
    for row in rows:
        if row.get("is_benchmark") or row.get("is_average") or row.get("error"):
            continue
        style = str(row.get("Style") or "").strip()
        value = row.get(period)
        if not style or value is None:
            continue
        grouped.setdefault(style, []).append(float(value))

    averages = [
        {"style": style, "count": len(values), "average": sum(values) / len(values)}
        for style, values in grouped.items()
        if values
    ]
    return sorted(averages, key=lambda row: (-float(row["average"]), str(row["style"])))


def _build_style_commentary(absolute_rows: list[dict], relative_rows: list[dict]) -> str:
    absolute_styles = _average_by_style(absolute_rows, "12M")
    relative_styles = _average_by_style(relative_rows, "12M")

    if not absolute_styles and not relative_styles:
        return "No 12M style averages are available yet."

    absolute_leader = absolute_styles[0] if absolute_styles else None
    absolute_trailer = absolute_styles[-1] if len(absolute_styles) > 1 else None
    relative_leader = relative_styles[0] if relative_styles else None

    if absolute_leader and relative_leader and absolute_leader["style"] == relative_leader["style"]:
        commentary = (
            f"On 12M style averages, {absolute_leader['style']} leads both total return "
            f"({_format_percent(float(absolute_leader['average']))}) and excess return "
            f"({_format_percent(float(relative_leader['average']))} versus the benchmark)."
        )
    elif absolute_leader and relative_leader:
        commentary = (
            f"On 12M style averages, {absolute_leader['style']} leads total return at "
            f"{_format_percent(float(absolute_leader['average']))}, while {relative_leader['style']} leads "
            f"excess return at {_format_percent(float(relative_leader['average']))} versus the benchmark."
        )
    elif absolute_leader:
        commentary = (
            f"On 12M style averages, {absolute_leader['style']} leads total return at "
            f"{_format_percent(float(absolute_leader['average']))}."
        )
    else:
        commentary = (
            f"On 12M style averages, {relative_leader['style']} leads excess return at "
            f"{_format_percent(float(relative_leader['average']))} versus the benchmark."
        )

    if absolute_trailer is not None and absolute_leader is not None and absolute_trailer["style"] != absolute_leader["style"]:
        commentary += (
            f" {absolute_trailer['style']} is the weakest total-return cohort at "
            f"{_format_percent(float(absolute_trailer['average']))}."
        )
    return commentary


def _report_snapshot(absolute_rows: list[dict], relative_rows: list[dict]) -> dict[str, object]:
    benchmark = next((row for row in absolute_rows if row.get("is_benchmark")), None)
    fund_rows = [row for row in absolute_rows if not row.get("is_benchmark") and not row.get("is_average")]
    fund_relative_rows = [row for row in relative_rows if not row.get("is_average")]
    valid_absolute = [row for row in fund_rows if not row.get("error") and row.get("12M") is not None]
    valid_relative = [row for row in fund_relative_rows if not row.get("error") and row.get("12M") is not None]

    return {
        "benchmark": benchmark,
        "fund_rows": fund_rows,
        "live_count": len(fund_rows),
        "stale_count": sum(1 for row in fund_rows if _row_is_stale(row)),
        "ahead_count": sum(1 for row in valid_relative if float(row["12M"]) > 0),
        "best_absolute": max(valid_absolute, key=lambda row: _safe_period_value(row, "12M"), default=None),
        "best_relative": max(valid_relative, key=lambda row: _safe_period_value(row, "12M"), default=None),
        "leaders": sorted(valid_absolute, key=lambda row: _safe_period_value(row, "12M"), reverse=True)[:5],
        "style_commentary": _build_style_commentary(fund_rows, fund_relative_rows),
    }


def _competitor_set_title(competitor_set) -> str:
    if isinstance(competitor_set, dict):
        return str(competitor_set.get("title") or "Competitor set")
    return str(getattr(competitor_set, "title", "Competitor set"))


def _competitor_set_id(competitor_set) -> str:
    if isinstance(competitor_set, dict):
        return str(competitor_set.get("id") or _competitor_set_title(competitor_set))
    return str(getattr(competitor_set, "id", _competitor_set_title(competitor_set)))


def _competitor_set_rows(competitor_set) -> list[dict]:
    if isinstance(competitor_set, dict):
        return list(competitor_set.get("rows") or [])
    return list(getattr(competitor_set, "rows", []) or [])


def render_tables(
    absolute_rows: list[dict],
    relative_rows: list[dict],
    as_of_date,
    console: Console | None = None,
    competitor_sets: list | None = None,
) -> None:
    console = console or Console()
    competitor_sets = competitor_sets or []
    competitor_rows = [row for competitor_set in competitor_sets for row in _competitor_set_rows(competitor_set)]
    if any(_row_is_stale(row) for row in absolute_rows + relative_rows + competitor_rows):
        console.print(
            "[yellow]Note: rows marked stale use the latest public fund date shown in the row label, not the report date in the title.[/yellow]"
        )
        console.print()
    console.print(build_rich_table(absolute_rows, f"Absolute Total Return Performance ({as_of_date:%Y-%m-%d})"))
    console.print()
    console.print(build_rich_table(relative_rows, f"Relative Total Return Performance ({as_of_date:%Y-%m-%d})"))
    for competitor_set in competitor_sets:
        console.print()
        console.print(
            build_rich_table(
                _competitor_set_rows(competitor_set),
                f"{_competitor_set_title(competitor_set)} ({as_of_date:%Y-%m-%d})",
                include_benchmark=False,
            )
        )


def build_rich_table(
    rows: list[dict],
    title: str,
    *,
    top_n: int = 3,
    bottom_n: int = 3,
    include_benchmark: bool = False,
) -> Table:
    table = Table(title=title)
    highlights = build_period_highlights(
        rows,
        periods=PERIODS,
        top_n=top_n,
        bottom_n=bottom_n,
        include_benchmark=include_benchmark,
    )
    table.add_column("Fund", style="bold", overflow="fold", min_width=24)
    table.add_column("Style")
    for period in PERIODS:
        label = f"{period} (p.a.)" if period in {"3Y", "5Y"} else period
        table.add_column(label, justify="right")

    for row_index, row in enumerate(rows):
        label = row["Fund"]
        if row.get("is_benchmark"):
            label = f"[bold]{label}[/bold]"
        if row.get("is_average"):
            label = f"[bold]{label}[/bold]"
        if _is_firetrail(row):
            label = f"[bold]{label}[/bold]"
        if row.get("is_disabled"):
            label = f"{label} [yellow]({row.get('disabled_reason') or 'Source pending'})[/yellow]"
        if _row_is_stale(row):
            latest_date = row.get("latest_date")
            latest_label = latest_date.strftime("%Y-%m-%d") if latest_date is not None else "latest public date"
            label = f"{label} [yellow](as of {latest_label}, stale {row['stale_days']}d)[/yellow]"

        values = [
            _format_value(row.get(period), error=row.get("error", False), highlight=highlights.get((row_index, period)))
            for period in PERIODS
        ]
        table.add_row(label, str(row.get("Style") or ""), *values)

    return table


def _build_plaintext_table(rows: list[dict]) -> str:
    headers = ["Fund", "Style", *[f"{period} p.a." if period in {"3Y", "5Y"} else period for period in PERIODS]]
    rendered_rows = [
        [
            _plain_row_label(row),
            str(row.get("Style") or ""),
            *[_format_percent(None if row.get("error") else row.get(period), error=row.get("error", False)) for period in PERIODS],
        ]
        for row in rows
    ]

    widths = [len(header) for header in headers]
    for row in rendered_rows:
        for index, value in enumerate(row):
            widths[index] = min(max(widths[index], len(value)), 42 if index == 0 else 10)

    def fit(value: str, width: int) -> str:
        return value if len(value) <= width else value[: max(width - 1, 1)] + "…"

    lines = [" | ".join(fit(header, widths[index]).ljust(widths[index]) for index, header in enumerate(headers))]
    lines.append("-|-".join("-" * width for width in widths))
    for row in rendered_rows:
        lines.append(" | ".join(fit(value, widths[index]).ljust(widths[index]) for index, value in enumerate(row)))
    return "\n".join(lines)


def build_plaintext_report(absolute_rows: list[dict], relative_rows: list[dict], as_of_date, competitor_sets: list | None = None) -> str:
    snapshot = _report_snapshot(absolute_rows, relative_rows)
    benchmark = snapshot["benchmark"]
    lines = [
        f"Australian Equity Fund Scorecard ({as_of_date:%Y-%m-%d})",
        "",
        "This report shows total returns, with distributions reinvested where needed.",
        "Relative performance is shown versus the S&P/ASX 200 Accumulation benchmark.",
    ]

    if benchmark is not None:
        lines.append(f"Benchmark 12M: {_format_percent(benchmark.get('12M'))}")
        lines.append(f"Benchmark 3Y p.a.: {_format_percent(benchmark.get('3Y'))}")

    lines.extend(
        [
            f"Funds ahead of benchmark over 12M: {snapshot['ahead_count']} of {snapshot['live_count']}",
            f"Funds with stale public data: {snapshot['stale_count']}",
        ]
    )

    best_absolute = snapshot["best_absolute"]
    best_relative = snapshot["best_relative"]
    if best_absolute is not None:
        lines.append(f"Best 12M total return: {best_absolute['Fund']} ({_format_percent(best_absolute.get('12M'))})")
    if best_relative is not None:
        lines.append(f"Best 12M excess return: {best_relative['Fund']} ({_format_percent(best_relative.get('12M'))})")
    lines.append(f"Style lens: {snapshot['style_commentary']}")

    leaders = snapshot["leaders"]
    if leaders:
        lines.extend(["", "Top 12M funds:"])
        for row in leaders:
            lines.append(f"- {_plain_row_label(row)}: {_format_percent(row.get('12M'))}")

    lines.extend(["", "The HTML version of this report includes the full styled tables."])
    for competitor_set in competitor_sets or []:
        lines.extend(["", f"{_competitor_set_title(competitor_set)}:", _build_plaintext_table(_competitor_set_rows(competitor_set))])
    return "\n".join(lines)


def build_html_report(absolute_rows: list[dict], relative_rows: list[dict], as_of_date, competitor_sets: list | None = None) -> str:
    snapshot = _report_snapshot(absolute_rows, relative_rows)
    benchmark = snapshot["benchmark"]
    best_absolute = snapshot["best_absolute"]
    best_relative = snapshot["best_relative"]

    summary_cards = [
        ("Live funds", str(snapshot["live_count"]), "Managers currently included in the daily report."),
        (
            "Funds ahead of benchmark (12M)",
            f"{snapshot['ahead_count']} / {snapshot['live_count']}",
            "How many live funds are ahead of the benchmark over the last 12 months.",
        ),
        (
            "Benchmark 12M",
            _format_percent(None if benchmark is None else benchmark.get("12M")),
            "S&P/ASX 200 Accumulation return over the last 12 months.",
        ),
        (
            "Stale sources",
            str(snapshot["stale_count"]),
            "Funds whose latest public data is older than the report date.",
        ),
    ]

    leaders_html = "".join(
        f"""
        <li>
          <span class="leader-name">{escape(str(row["Fund"]))}</span>
          <span class="leader-value">{escape(_format_percent(row.get("12M")))}</span>
        </li>
        """
        for row in snapshot["leaders"]
    )
    if not leaders_html:
        leaders_html = "<li><span class=\"leader-name\">No 12M leaders available</span><span class=\"leader-value\">N/A</span></li>"

    benchmark_blurb = (
        f"{escape(_format_percent(benchmark.get('MTD')))} MTD and {escape(_format_percent(benchmark.get('12M')))} over 12 months."
        if benchmark is not None
        else "Benchmark data unavailable."
    )
    best_absolute_blurb = (
        f"{escape(str(best_absolute['Fund']))} is the strongest 12M total-return result at {escape(_format_percent(best_absolute.get('12M')))}."
        if best_absolute is not None
        else "No 12M absolute leader is available."
    )
    best_relative_blurb = (
        f"{escape(str(best_relative['Fund']))} leads on 12M excess return at {escape(_format_percent(best_relative.get('12M')))}."
        if best_relative is not None
        else "No 12M relative leader is available."
    )

    cards_html = "".join(
        f"""
        <article class="stat-card">
          <p class="stat-label">{escape(label)}</p>
          <p class="stat-value">{escape(value)}</p>
          <p class="stat-note">{escape(note)}</p>
        </article>
        """
        for label, value, note in summary_cards
    )
    competitor_sections_html = _build_competitor_set_sections_html(competitor_sets or [])

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Australian Equity Fund Scorecard | {as_of_date:%Y-%m-%d}</title>
  <style>
    :root {{
      color-scheme: light;
      --bg-top: #f2ebe0;
      --bg-bottom: #eef5ef;
      --panel: rgba(255, 255, 255, 0.9);
      --panel-strong: rgba(255, 255, 255, 0.96);
      --ink: #17352f;
      --muted: #5d6b66;
      --border: #d5dccf;
      --accent: #1f6a5b;
      --accent-soft: #dbece7;
      --warm: #bb6d43;
      --warm-soft: #f0dfd4;
      --green: #1f7d4b;
      --green-soft: #def1e5;
      --red: #b24f3e;
      --red-soft: #f6dfd9;
      --neutral: #58656a;
      --neutral-soft: #edf1f2;
      --shadow: 0 18px 40px rgba(25, 47, 42, 0.12);
      --radius-lg: 26px;
      --radius-md: 18px;
      --radius-sm: 999px;
    }}
    * {{
      box-sizing: border-box;
    }}
    body {{
      margin: 0;
      font-family: "Aptos", "Segoe UI", "Helvetica Neue", Arial, sans-serif;
      color: var(--ink);
      background:
        radial-gradient(circle at top left, rgba(187, 109, 67, 0.16), transparent 35%),
        radial-gradient(circle at top right, rgba(31, 106, 91, 0.18), transparent 35%),
        linear-gradient(180deg, var(--bg-top), var(--bg-bottom));
    }}
    .shell {{
      width: min(1200px, calc(100% - 32px));
      margin: 0 auto;
      padding: 32px 0 40px;
    }}
    .hero {{
      padding: 32px;
      border: 1px solid rgba(255, 255, 255, 0.55);
      border-radius: var(--radius-lg);
      background: linear-gradient(135deg, rgba(21, 54, 47, 0.92), rgba(31, 106, 91, 0.88));
      color: #f7f3eb;
      box-shadow: var(--shadow);
      overflow: hidden;
      position: relative;
    }}
    .hero::after {{
      content: "";
      position: absolute;
      inset: auto -10% -35% auto;
      width: 320px;
      height: 320px;
      border-radius: 50%;
      background: rgba(255, 255, 255, 0.08);
    }}
    .eyebrow {{
      margin: 0 0 10px;
      text-transform: uppercase;
      letter-spacing: 0.18em;
      font-size: 12px;
      opacity: 0.78;
    }}
    h1 {{
      margin: 0;
      font-size: clamp(34px, 5vw, 54px);
      line-height: 1.02;
      max-width: 12ch;
    }}
    .lede {{
      margin: 16px 0 0;
      max-width: 760px;
      font-size: 17px;
      line-height: 1.6;
      color: rgba(247, 243, 235, 0.9);
    }}
    .hero-meta {{
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
      margin-top: 22px;
    }}
    .badge {{
      display: inline-flex;
      align-items: center;
      padding: 10px 14px;
      border-radius: var(--radius-sm);
      background: rgba(255, 255, 255, 0.12);
      border: 1px solid rgba(255, 255, 255, 0.18);
      font-size: 14px;
    }}
    .stats {{
      display: grid;
      grid-template-columns: repeat(4, minmax(0, 1fr));
      gap: 16px;
      margin-top: 22px;
    }}
    .stat-card {{
      padding: 22px;
      border-radius: var(--radius-md);
      background: var(--panel);
      border: 1px solid rgba(255, 255, 255, 0.6);
      box-shadow: var(--shadow);
      backdrop-filter: blur(6px);
    }}
    .stat-label {{
      margin: 0;
      color: var(--muted);
      text-transform: uppercase;
      letter-spacing: 0.08em;
      font-size: 12px;
    }}
    .stat-value {{
      margin: 10px 0 8px;
      font-size: 34px;
      font-weight: 700;
      line-height: 1.05;
    }}
    .stat-note {{
      margin: 0;
      color: var(--muted);
      line-height: 1.5;
      font-size: 14px;
    }}
    .insights {{
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 16px;
      margin-top: 16px;
    }}
    .insight-card,
    .table-card {{
      padding: 24px;
      border-radius: var(--radius-lg);
      background: var(--panel-strong);
      border: 1px solid rgba(255, 255, 255, 0.65);
      box-shadow: var(--shadow);
    }}
    .insight-kicker {{
      margin: 0 0 10px;
      color: var(--warm);
      text-transform: uppercase;
      letter-spacing: 0.1em;
      font-size: 12px;
    }}
    .insight-card h2,
    .table-card h2 {{
      margin: 0;
      font-size: 24px;
      line-height: 1.2;
    }}
    .insight-card p,
    .table-card p {{
      margin: 10px 0 0;
      color: var(--muted);
      line-height: 1.6;
    }}
    .leader-list {{
      list-style: none;
      margin: 18px 0 0;
      padding: 0;
      display: grid;
      gap: 10px;
    }}
    .leader-list li {{
      display: flex;
      justify-content: space-between;
      align-items: center;
      gap: 16px;
      padding: 12px 14px;
      border-radius: 16px;
      background: #f8faf7;
      border: 1px solid var(--border);
    }}
    .leader-name {{
      font-weight: 600;
    }}
    .leader-value {{
      white-space: nowrap;
      color: var(--accent);
      font-weight: 700;
    }}
    .table-card {{
      margin-top: 16px;
    }}
    .table-wrap {{
      margin-top: 18px;
      overflow-x: auto;
      border-radius: 18px;
      border: 1px solid var(--border);
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      min-width: 760px;
      background: white;
    }}
    th,
    td {{
      padding: 14px 16px;
      text-align: left;
      border-bottom: 1px solid #eef1ec;
      vertical-align: top;
    }}
    th {{
      position: sticky;
      top: 0;
      z-index: 1;
      background: #f5f7f2;
      color: var(--muted);
      font-size: 12px;
      letter-spacing: 0.08em;
      text-transform: uppercase;
    }}
    th:first-child,
    td:first-child {{
      position: sticky;
      left: 0;
      z-index: 1;
      background: inherit;
    }}
    tbody tr:nth-child(odd) {{
      background: #fbfcfa;
    }}
    tbody tr.benchmark-row {{
      background: #f0f6f2;
    }}
    tbody tr.benchmark-row td:first-child {{
      background: #f0f6f2;
    }}
    tbody tr.summary-row {{
      background: #eef5ef;
    }}
    tbody tr.summary-row td {{
      font-weight: 700;
    }}
    tbody tr.summary-row td:first-child {{
      background: #eef5ef;
    }}
    tbody tr:hover {{
      background: #f4f8f6;
    }}
    tbody tr:hover td:first-child {{
      background: #f4f8f6;
    }}
    .fund-name {{
      font-weight: 700;
      color: var(--ink);
    }}
    .fund-meta {{
      margin-top: 6px;
      color: var(--muted);
      font-size: 13px;
    }}
    .pill {{
      display: inline-flex;
      align-items: center;
      justify-content: center;
      min-width: 76px;
      padding: 8px 10px;
      border-radius: var(--radius-sm);
      font-weight: 700;
      font-size: 14px;
      white-space: nowrap;
    }}
    .pill.positive {{
      color: var(--green);
      background: var(--green-soft);
    }}
    .pill.negative {{
      color: var(--red);
      background: var(--red-soft);
    }}
    .pill.top-performer {{
      color: var(--green);
      background: var(--green-soft);
      box-shadow: inset 0 0 0 2px rgba(31, 125, 75, 0.28);
    }}
    .pill.bottom-performer {{
      color: var(--red);
      background: var(--red-soft);
      box-shadow: inset 0 0 0 2px rgba(178, 79, 62, 0.28);
    }}
    .pill.neutral {{
      color: var(--neutral);
      background: var(--neutral-soft);
    }}
    .pill.error {{
      color: var(--red);
      background: var(--red-soft);
    }}
    .style-label {{
      color: var(--muted);
      font-weight: 600;
    }}
    .footer {{
      margin-top: 16px;
      padding: 18px 24px;
      border-radius: 20px;
      background: rgba(255, 255, 255, 0.72);
      border: 1px solid rgba(255, 255, 255, 0.7);
      color: var(--muted);
      line-height: 1.6;
      font-size: 14px;
    }}
    @media (max-width: 960px) {{
      .stats,
      .insights {{
        grid-template-columns: 1fr 1fr;
      }}
    }}
    @media (max-width: 680px) {{
      .shell {{
        width: min(100% - 18px, 1200px);
        padding-top: 18px;
      }}
      .hero,
      .stat-card,
      .insight-card,
      .table-card {{
        padding: 18px;
      }}
      .stats,
      .insights {{
        grid-template-columns: 1fr;
      }}
      th:first-child,
      td:first-child {{
        position: static;
      }}
    }}
  </style>
</head>
<body>
  <main class="shell">
    <section class="hero">
      <p class="eyebrow">Daily distribution-aware performance snapshot</p>
      <h1>Australian Equity Fund Scorecard</h1>
      <p class="lede">
        A cleaner, more approachable view of manager performance for non-technical readers.
        All figures are total return, and relative numbers are shown against the S&amp;P/ASX 200 Accumulation benchmark.
      </p>
      <div class="hero-meta">
        <span class="badge">As of {as_of_date:%Y-%m-%d}</span>
        <span class="badge">{snapshot['live_count']} live funds</span>
        <span class="badge">{snapshot['stale_count']} stale public sources</span>
      </div>
    </section>

    <section class="stats">
      {cards_html}
    </section>

    <section class="insights">
      <article class="insight-card">
        <p class="insight-kicker">Benchmark pulse</p>
        <h2>S&amp;P/ASX 200 Accumulation</h2>
        <p>{benchmark_blurb}</p>
      </article>
      <article class="insight-card">
        <p class="insight-kicker">Strongest total return</p>
        <h2>Best 12M performer</h2>
        <p>{best_absolute_blurb}</p>
      </article>
      <article class="insight-card">
        <p class="insight-kicker">Strongest excess return</p>
        <h2>Best 12M versus benchmark</h2>
        <p>{best_relative_blurb}</p>
      </article>
    </section>

    <section class="table-card">
      <p class="insight-kicker">Style lens</p>
      <h2>How styles are tracking</h2>
      <p>{escape(str(snapshot['style_commentary']))}</p>
    </section>

    <section class="table-card">
      <p class="insight-kicker">Leaderboard</p>
      <h2>Top 12M funds</h2>
      <p>The strongest trailing-12-month total-return results among live managers.</p>
      <ul class="leader-list">
        {leaders_html}
      </ul>
    </section>

    <section class="table-card">
      <p class="insight-kicker">Absolute performance</p>
      <h2>Total return table</h2>
      <p>Returns shown after reinvesting distributions where required. Green highlights mark the best results per period; red highlights mark the worst results.</p>
      <div class="table-wrap">
        {_build_html_table(absolute_rows, include_benchmark_marker=True)}
      </div>
    </section>

    <section class="table-card">
      <p class="insight-kicker">Relative performance</p>
      <h2>Return above or below benchmark</h2>
      <p>Relative performance shows each fund's excess return against the S&amp;P/ASX 200 Accumulation index.</p>
      <div class="table-wrap">
        {_build_html_table(relative_rows, include_benchmark_marker=False)}
      </div>
    </section>

    {competitor_sections_html}

    <section class="footer">
      Rows marked stale use the latest public fund date shown in the row, not the headline report date.
      Average rows are the simple mean of live fund returns excluding the benchmark. Three-year and five-year figures are annualized.
    </section>
  </main>
</body>
</html>"""


def _build_competitor_set_sections_html(competitor_sets: list) -> str:
    sections = []
    for competitor_set in competitor_sets:
        title = _competitor_set_title(competitor_set)
        rows = _competitor_set_rows(competitor_set)
        sections.append(
            f"""
    <section class="table-card" id="{escape(_competitor_set_id(competitor_set))}">
      <p class="insight-kicker">Competitor set</p>
      <h2>{escape(title)}</h2>
      <div class="table-wrap">
        {_build_html_table(rows, include_benchmark_marker=True, include_benchmark_highlight=False)}
      </div>
    </section>
            """
        )
    return "".join(sections)


def _build_html_table(
    rows: list[dict],
    include_benchmark_marker: bool,
    *,
    top_n: int = 3,
    bottom_n: int = 3,
    include_benchmark_highlight: bool = False,
) -> str:
    highlights = build_period_highlights(
        rows,
        periods=PERIODS,
        top_n=top_n,
        bottom_n=bottom_n,
        include_benchmark=include_benchmark_highlight,
    )
    header_cells = "".join(
        f"<th>{escape(f'{period} (p.a.)' if period in {'3Y', '5Y'} else period)}</th>" for period in PERIODS
    )

    body_rows = []
    for row_index, row in enumerate(rows):
        row_classes: list[str] = []
        if include_benchmark_marker and row.get("is_benchmark"):
            row_classes.append("benchmark-row")
        if row.get("is_average"):
            row_classes.append("summary-row")
        if _is_firetrail(row):
            row_classes.append("firetrail-row")
        latest_date = row.get("latest_date")
        stale_meta = ""
        if _row_is_stale(row):
            stale_meta = (
                f"<div class=\"fund-meta\">Public data through {_format_public_date(latest_date)} "
                f"({int(row['stale_days'])} day{'s' if int(row['stale_days']) != 1 else ''} stale)</div>"
            )
        elif row.get("is_disabled"):
            stale_meta = f"<div class=\"fund-meta\">{escape(str(row.get('disabled_reason') or 'Source pending'))}</div>"
        elif latest_date is not None:
            stale_meta = f"<div class=\"fund-meta\">Public data through {_format_public_date(latest_date)}</div>"

        cells = "".join(
            _build_html_value_cell(row.get(period), error=row.get("error", False), highlight=highlights.get((row_index, period)))
            for period in PERIODS
        )
        body_rows.append(
            f"""
            <tr class="{' '.join(row_classes)}">
              <td>
                <div class="fund-name">{escape(str(row["Fund"]))}</div>
                {stale_meta}
              </td>
              <td><span class="style-label">{escape(str(row.get("Style") or ""))}</span></td>
              {cells}
            </tr>
            """
        )

    return f"""
    <table>
      <thead>
        <tr>
          <th>Fund</th>
          <th>Style</th>
          {header_cells}
        </tr>
      </thead>
      <tbody>
        {''.join(body_rows)}
      </tbody>
    </table>
    """


def _build_html_value_cell(value: float | None, error: bool = False, highlight: PerformanceHighlight | None = None) -> str:
    classes = ["pill"]
    if error:
        classes.append("error")
    elif value is None:
        classes.append("neutral")
    elif highlight == HIGHLIGHT_TOP:
        classes.append("top-performer")
    elif highlight == HIGHLIGHT_BOTTOM:
        classes.append("bottom-performer")
    elif value >= 0:
        classes.append("positive")
    else:
        classes.append("negative")

    label = escape(_format_percent(value, error=error))
    return f"<td><span class=\"{' '.join(classes)}\">{label}</span></td>"


def _sheet_name(base: str, used_names: set[str]) -> str:
    safe = "".join(char if char not in r"[]:*?/\\" else " " for char in base).strip() or "Sheet"
    safe = safe[:31]
    candidate = safe
    suffix = 2
    while candidate in used_names:
        tail = f" {suffix}"
        candidate = f"{safe[:31 - len(tail)]}{tail}"
        suffix += 1
    used_names.add(candidate)
    return candidate


def export_to_excel(
    absolute_rows: list[dict],
    relative_rows: list[dict],
    as_of_date,
    output_path: str | Path,
    competitor_sets: list | None = None,
) -> Path:
    output = Path(output_path)
    absolute_df = _rows_to_dataframe(absolute_rows)
    relative_df = _rows_to_dataframe(relative_rows)
    rows_by_sheet = {"Absolute": absolute_rows, "Relative": relative_rows}

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        absolute_df.to_excel(writer, sheet_name="Absolute", index=False)
        relative_df.to_excel(writer, sheet_name="Relative", index=False)
        used_sheet_names = {"Absolute", "Relative"}
        for competitor_set in competitor_sets or []:
            sheet_name = _sheet_name(_competitor_set_title(competitor_set), used_sheet_names)
            rows = _competitor_set_rows(competitor_set)
            _rows_to_dataframe(rows).to_excel(writer, sheet_name=sheet_name, index=False)
            rows_by_sheet[sheet_name] = rows

    workbook = load_workbook(output)
    top_fill = PatternFill(fill_type="solid", fgColor="FFDEF1E5")
    bottom_fill = PatternFill(fill_type="solid", fgColor="FFF6DFD9")
    for sheet_name in rows_by_sheet:
        highlights = build_period_highlights(
            rows_by_sheet[sheet_name],
            periods=PERIODS,
            top_n=3,
            bottom_n=3,
            include_benchmark=False,
        )
        worksheet = workbook[sheet_name]
        worksheet["A1"].font = Font(bold=True)
        header_to_column = {worksheet.cell(row=1, column=index).value: index for index in range(1, worksheet.max_column + 1)}
        period_start_column = header_to_column.get(PERIODS[0], worksheet.max_column)

        for data_row_index, row in enumerate(worksheet.iter_rows(min_row=2), start=0):
            source_row = rows_by_sheet[sheet_name][data_row_index]
            if _is_firetrail(source_row):
                row[0].font = Font(bold=True)
            for cell in row[period_start_column - 1 :]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.0%"
                    header = worksheet.cell(row=1, column=cell.column).value
                    highlight = highlights.get((data_row_index, str(header)))
                    if highlight == HIGHLIGHT_TOP:
                        cell.font = Font(color="008000", bold=True)
                        cell.fill = top_fill
                    elif highlight == HIGHLIGHT_BOTTOM:
                        cell.font = Font(color="9C0006", bold=True)
                        cell.fill = bottom_fill
                    elif cell.value >= 0:
                        cell.font = Font(color="008000")
                    else:
                        cell.font = Font(color="9C0006")

        for column in worksheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            worksheet.column_dimensions[column[0].column_letter].width = max_length + 2

    workbook.save(output)
    return output


def export_to_html(
    absolute_rows: list[dict],
    relative_rows: list[dict],
    as_of_date,
    output_path: str | Path,
    competitor_sets: list | None = None,
) -> Path:
    output = Path(output_path)
    output.write_text(build_html_report(absolute_rows, relative_rows, as_of_date, competitor_sets=competitor_sets), encoding="utf-8")
    return output


def _rows_to_dataframe(rows: list[dict]):
    records = []
    for row in rows:
        latest_date = row.get("latest_date")
        record = {
            "Fund": row["Fund"],
            "Style": row.get("Style") or "",
            "As Of": None if latest_date is None else str(pd.Timestamp(latest_date).date()),
        }
        for period in PERIODS:
            record[period] = None if row.get("error") else row.get(period)
        record["Status"] = row.get("disabled_reason") if row.get("is_disabled") else ("Error" if row.get("error") else "")
        records.append(record)
    return pd.DataFrame(records)
