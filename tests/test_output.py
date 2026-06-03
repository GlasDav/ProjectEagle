from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from rich.console import Console

from output import _build_html_table, build_html_report, build_plaintext_report, build_rich_table, export_to_excel


def _ranked_rows(count: int = 7):
    rows = []
    for index in range(1, count + 1):
        rows.append(
            {
                "Fund": f"Fund {index}",
                "Style": "Growth",
                "is_benchmark": False,
                "is_stale": False,
                "error": False,
                "stale_days": 0,
                "latest_date": None,
                "MTD": index / 100,
                "3M": index / 100,
                "6M": index / 100,
                "12M": index / 100,
                "3Y": index / 100,
                "5Y": index / 100,
            }
        )
    return rows


def _firetrail_rows():
    rows = _ranked_rows(3)
    rows[0]["Fund"] = "Firetrail High Conviction Fund"
    rows[1]["Fund"] = "Firetrail Alpha Plus Fund Complex ETF"
    rows[2]["Fund"] = "Other Fund"
    return rows


def _ranked_rows_with_high_benchmark():
    benchmark = {
        "Fund": "S&P/ASX 200 Accumulation",
        "Style": "",
        "is_benchmark": True,
        "is_stale": False,
        "error": False,
        "stale_days": 0,
        "latest_date": None,
        "MTD": 0.99,
        "3M": 0.99,
        "6M": 0.99,
        "12M": 0.99,
        "3Y": 0.99,
        "5Y": 0.99,
    }
    return [benchmark, *_ranked_rows()]


def _sample_rows():
    absolute_rows = [
        {
            "Fund": "Benchmark",
            "Style": "",
            "is_benchmark": True,
            "is_stale": False,
            "error": False,
            "stale_days": 0,
            "latest_date": None,
            "MTD": 0.01,
            "3M": 0.02,
            "6M": 0.03,
            "12M": 0.04,
            "3Y": 0.05,
            "5Y": 0.06,
        },
        {
            "Fund": "Fund A",
            "Style": "Growth",
            "is_benchmark": False,
            "is_stale": False,
            "error": False,
            "stale_days": 0,
            "latest_date": None,
            "MTD": 0.02,
            "3M": 0.03,
            "6M": 0.04,
            "12M": 0.05,
            "3Y": 0.06,
            "5Y": 0.07,
        },
    ]
    relative_rows = [
        {
            "Fund": "Fund A",
            "Style": "Growth",
            "is_stale": False,
            "error": False,
            "stale_days": 0,
            "latest_date": None,
            "MTD": 0.01,
            "3M": 0.01,
            "6M": 0.01,
            "12M": 0.01,
            "3Y": 0.01,
            "5Y": 0.01,
        }
    ]
    return absolute_rows, relative_rows


def _competitor_sets():
    return [
        {
            "id": "long_short_funds",
            "title": "Long-short funds",
            "rows": [
                {
                    "Fund": "S&P/ASX 200 Accumulation",
                    "Style": "",
                    "is_benchmark": True,
                    "is_stale": False,
                    "error": False,
                    "stale_days": 0,
                    "latest_date": None,
                    "MTD": 0.01,
                    "3M": 0.02,
                    "6M": 0.03,
                    "12M": 0.04,
                    "3Y": 0.05,
                    "5Y": 0.06,
                },
                {
                    "Fund": "Sage Capital Equity Plus Fund",
                    "Style": "Agnostic",
                    "is_disabled": True,
                    "disabled_reason": "No durable public historical unit price and distribution feed has been validated.",
                    "is_stale": False,
                    "error": False,
                    "stale_days": 0,
                    "latest_date": None,
                    "MTD": None,
                    "3M": None,
                    "6M": None,
                    "12M": None,
                    "3Y": None,
                    "5Y": None,
                },
            ],
        }
    ]


def test_build_rich_table_includes_style_column():
    absolute_rows, _ = _sample_rows()

    table = build_rich_table(absolute_rows, "Example")

    assert [column.header for column in table.columns] == ["Fund", "Style", "MTD", "3M", "6M", "12M", "3Y (p.a.)", "5Y (p.a.)"]


def test_build_rich_table_wraps_stale_as_of_date_without_ellipsis():
    rows = [
        {
            "Fund": "Sage Capital Absolute Return Fund",
            "Style": "Agnostic",
            "is_benchmark": False,
            "is_stale": True,
            "error": False,
            "stale_days": 9,
            "latest_date": pd.Timestamp("2026-05-18"),
            "MTD": 0.059,
            "3M": 0.116,
            "6M": 0.017,
            "12M": -0.065,
            "3Y": -0.08,
            "5Y": -0.029,
        }
    ]
    console = Console(record=True, width=90, force_terminal=False, color_system=None)

    console.print(build_rich_table(rows, "Example"))
    rendered = console.export_text()

    assert "as of" in rendered
    assert "2026-05-18" in rendered
    assert "stale" in rendered
    assert "…" not in rendered


def test_plaintext_report_labels_non_stale_date_offsets_without_stale_count():
    absolute_rows, relative_rows = _sample_rows()
    absolute_rows[1]["latest_date"] = pd.Timestamp("2026-03-28")
    absolute_rows[1]["stale_days"] = 1
    absolute_rows[1]["is_stale"] = False

    report = build_plaintext_report(absolute_rows, relative_rows, pd.Timestamp("2026-03-29"))

    assert "Fund A (as of 2026-03-28)" in report
    assert "Funds with stale public data: 0" in report

def test_build_html_report_includes_style_column_and_values():
    absolute_rows, relative_rows = _sample_rows()

    html = build_html_report(absolute_rows, relative_rows, pd.Timestamp("2026-03-29"))

    assert "<th>Style</th>" in html
    assert ">Growth<" in html
    assert "How styles are tracking" in html


def test_build_html_table_uses_dynamic_top_and_bottom_highlight_count():
    html = _build_html_table(_ranked_rows(), include_benchmark_marker=False)

    assert html.count('class="pill top-performer"') == 12
    assert html.count('class="pill bottom-performer"') == 12


def test_build_html_report_competitor_sets_use_dynamic_top_and_bottom_highlight_count():
    absolute_rows, relative_rows = _sample_rows()
    competitor_sets = [{"id": "competitors", "title": "Competitors", "rows": _ranked_rows()}]

    html = build_html_report(absolute_rows, relative_rows, pd.Timestamp("2026-03-29"), competitor_sets=competitor_sets)
    competitor_html = html.split('<section class="table-card" id="competitors">', 1)[1]

    assert competitor_html.count('class="pill top-performer"') == 12
    assert competitor_html.count('class="pill bottom-performer"') == 12


def test_build_html_report_appends_competitor_set_tables_without_commentary():
    absolute_rows, relative_rows = _sample_rows()

    html = build_html_report(absolute_rows, relative_rows, pd.Timestamp("2026-03-29"), competitor_sets=_competitor_sets())

    assert "Long-short funds" in html
    assert "Sage Capital Equity Plus Fund" in html
    assert "No durable public historical unit price" in html


def test_build_html_report_adds_print_page_groups():
    absolute_rows, relative_rows = _sample_rows()

    html = build_html_report(absolute_rows, relative_rows, pd.Timestamp("2026-03-29"), competitor_sets=_competitor_sets())

    assert "@media print" in html
    assert "@page" in html
    assert 'class="print-page scorecard-page"' in html
    assert 'class="table-card print-page absolute-page"' in html
    assert 'class="table-card print-page relative-page"' in html
    assert 'class="print-page competitor-page last-print-page"' in html


def test_build_plaintext_report_uses_threshold_aware_stale_count():
    absolute_rows, relative_rows = _sample_rows()
    absolute_rows[1]["latest_date"] = pd.Timestamp("2026-03-20")
    absolute_rows[1]["stale_days"] = 9
    absolute_rows[1]["is_stale"] = False
    relative_rows[0]["latest_date"] = pd.Timestamp("2026-03-20")
    relative_rows[0]["stale_days"] = 9
    relative_rows[0]["is_stale"] = False

    report = build_plaintext_report(absolute_rows, relative_rows, pd.Timestamp("2026-03-29"))

    assert "Funds with stale public data: 0" in report
    assert "Style lens:" in report


def test_build_html_report_renders_average_row():
    absolute_rows, relative_rows = _sample_rows()
    absolute_rows.append(
        {
            "Fund": "Average",
            "Style": "",
            "is_average": True,
            "is_stale": False,
            "error": False,
            "stale_days": 0,
            "latest_date": None,
            "MTD": 0.02,
            "3M": 0.03,
            "6M": 0.04,
            "12M": 0.05,
            "3Y": 0.06,
            "5Y": 0.07,
        }
    )
    relative_rows.append(
        {
            "Fund": "Average",
            "Style": "",
            "is_average": True,
            "is_stale": False,
            "error": False,
            "stale_days": 0,
            "latest_date": None,
            "MTD": 0.01,
            "3M": 0.01,
            "6M": 0.01,
            "12M": 0.01,
            "3Y": 0.01,
            "5Y": 0.01,
        }
    )

    html = build_html_report(absolute_rows, relative_rows, pd.Timestamp("2026-03-29"))

    assert "Average" in html
    assert "Average rows are the simple mean of live fund returns excluding the benchmark." in html


def test_export_to_excel_writes_style_column(tmp_path: Path):
    absolute_rows, relative_rows = _sample_rows()
    output_path = tmp_path / "report.xlsx"

    export_to_excel(absolute_rows, relative_rows, pd.Timestamp("2026-03-29"), output_path)

    workbook = load_workbook(output_path)
    absolute_sheet = workbook["Absolute"]
    assert absolute_sheet["A1"].value == "Fund"
    assert absolute_sheet["B1"].value == "Style"
    assert absolute_sheet["B2"].value is None
    assert absolute_sheet["B3"].value == "Growth"


def test_export_to_excel_uses_dynamic_top_and_bottom_highlight_count(tmp_path: Path):
    rows = _ranked_rows()
    output_path = tmp_path / "report.xlsx"

    export_to_excel(rows, rows, pd.Timestamp("2026-03-29"), output_path)

    workbook = load_workbook(output_path)
    sheet = workbook["Absolute"]

    assert sheet["D2"].font.bold is True
    assert sheet["D2"].font.color.rgb == "009C0006"
    assert sheet["D2"].fill.fgColor.rgb == "FFF6DFD9"
    assert sheet["D4"].font.bold is False
    assert sheet["D8"].font.bold is True
    assert sheet["D8"].font.color.rgb == "00008000"
    assert sheet["D8"].fill.fgColor.rgb == "FFDEF1E5"


def test_export_to_excel_competitor_sets_use_dynamic_top_and_bottom_highlight_count(tmp_path: Path):
    absolute_rows, relative_rows = _sample_rows()
    output_path = tmp_path / "report.xlsx"

    export_to_excel(
        absolute_rows,
        relative_rows,
        pd.Timestamp("2026-03-29"),
        output_path,
        competitor_sets=[{"id": "competitors", "title": "Competitors", "rows": _ranked_rows()}],
    )

    workbook = load_workbook(output_path)
    sheet = workbook["Competitors"]

    assert sheet["D2"].font.bold is True
    assert sheet["D2"].fill.fgColor.rgb == "FFF6DFD9"
    assert sheet["D3"].font.bold is True
    assert sheet["D3"].fill.fgColor.rgb == "FFF6DFD9"
    assert sheet["D4"].font.bold is False
    assert sheet["D5"].font.bold is False
    assert sheet["D6"].font.bold is False
    assert sheet["D7"].font.bold is True
    assert sheet["D7"].fill.fgColor.rgb == "FFDEF1E5"
    assert sheet["D8"].font.bold is True
    assert sheet["D8"].fill.fgColor.rgb == "FFDEF1E5"


def test_export_to_excel_competitor_set_highlights_ignore_benchmark_row(tmp_path: Path):
    absolute_rows, relative_rows = _sample_rows()
    output_path = tmp_path / "report.xlsx"

    export_to_excel(
        absolute_rows,
        relative_rows,
        pd.Timestamp("2026-03-29"),
        output_path,
        competitor_sets=[{"id": "competitors", "title": "Competitors", "rows": _ranked_rows_with_high_benchmark()}],
    )

    workbook = load_workbook(output_path)
    sheet = workbook["Competitors"]

    assert sheet["D2"].value == 0.99
    assert sheet["D2"].font.bold is False
    assert sheet["D2"].fill.fgColor.rgb == "00000000"
    assert sheet["D3"].fill.fgColor.rgb == "FFF6DFD9"
    assert sheet["D4"].fill.fgColor.rgb == "FFF6DFD9"
    assert sheet["D5"].fill.fgColor.rgb == "00000000"
    assert sheet["D6"].fill.fgColor.rgb == "00000000"
    assert sheet["D7"].fill.fgColor.rgb == "00000000"
    assert sheet["D8"].fill.fgColor.rgb == "FFDEF1E5"
    assert sheet["D9"].fill.fgColor.rgb == "FFDEF1E5"


def test_export_to_excel_bolds_all_firetrail_fund_rows(tmp_path: Path):
    rows = _firetrail_rows()
    output_path = tmp_path / "report.xlsx"

    export_to_excel(rows, rows, pd.Timestamp("2026-03-29"), output_path)

    workbook = load_workbook(output_path)
    sheet = workbook["Absolute"]

    assert sheet["A2"].font.bold is True
    assert sheet["A3"].font.bold is True
    assert sheet["A4"].font.bold is False


def test_export_to_excel_appends_competitor_set_sheets(tmp_path: Path):
    absolute_rows, relative_rows = _sample_rows()
    output_path = tmp_path / "report.xlsx"

    export_to_excel(
        absolute_rows,
        relative_rows,
        pd.Timestamp("2026-03-29"),
        output_path,
        competitor_sets=_competitor_sets(),
    )

    workbook = load_workbook(output_path)
    sheet = workbook["Long-short funds"]

    assert sheet["A2"].value == "S&P/ASX 200 Accumulation"
    assert sheet["A3"].value == "Sage Capital Equity Plus Fund"
    assert "No durable public" in sheet["J3"].value


def test_build_plaintext_report_appends_competitor_set_table():
    absolute_rows, relative_rows = _sample_rows()

    report = build_plaintext_report(absolute_rows, relative_rows, pd.Timestamp("2026-03-29"), competitor_sets=_competitor_sets())

    assert "Long-short funds:" in report
    assert "Sage Capital Equity Plus Fund" in report
